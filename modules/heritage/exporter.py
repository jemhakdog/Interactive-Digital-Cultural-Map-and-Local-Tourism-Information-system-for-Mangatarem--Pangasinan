"""
NCCA Form Exporter — generates perfectly formatted Excel sheets
for cultural heritage registries 01-07.
"""

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime

def export_heritage_excel(profile, config):
    """
    Generates a beautifully styled NCCA-compliant Excel file from a HeritageProfile.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NCCA Registry Form"
    ws.views.sheetView[0].showGridLines = True

    # ---- Styling Palette ----
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="1E293B")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="64748B")
    section_font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
    label_font = Font(name=font_family, size=10, bold=True, color="334155")
    value_font = Font(name=font_family, size=10, color="000000")
    
    # Fills (Harmony Colors - Teal/Slate theme)
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid") # Dark Teal
    label_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Light Gray/Slate
    value_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Borders
    thin_border = Side(style='thin', color="CBD5E1")
    cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    thick_bottom = Border(bottom=Side(style='medium', color="0F766E"))

    # ---- Headers ----
    ws.merge_cells('A1:B1')
    ws['A1'] = "Interactive Digital Cultural Map of Mangatarem, Pangasinan"
    ws['A1'].font = subtitle_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:B2')
    ws['A2'] = f"NCCA REGISTRY FORM {config['form']} ({config['label'].upper()})"
    ws['A2'].font = title_font
    ws['A2'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 25
    
    ws.merge_cells('A3:B3')
    ws['A3'] = "Municipal Local Tourism Office - Compliance Record"
    ws['A3'].font = subtitle_font
    ws['A3'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 18

    # Add a thin spacing row with border
    ws['A4'].border = thick_bottom
    ws['B4'].border = thick_bottom
    ws.row_dimensions[4].height = 10

    row_idx = 6

    # ---- Meta Fields (Base Attributes) ----
    meta_fields = [
        ("Status", profile.status.upper()),
        ("Date Profiled", profile.date_profiled.strftime("%B %d, %Y") if profile.date_profiled else "TBD"),
        ("Mapper Name", profile.mapper_name or "TBD"),
        ("System ID", f"MNG-HER-{profile.id:04d}"),
    ]
    
    # Section Header: Metadata
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    sec_cell = ws.cell(row=row_idx, column=1)
    sec_cell.value = "ADMINISTRATIVE & MAPPER METADATA"
    sec_cell.font = section_font
    sec_cell.fill = header_fill
    sec_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_idx].height = 24
    row_idx += 1

    for label, val in meta_fields:
        # Label cell
        l_cell = ws.cell(row=row_idx, column=1)
        l_cell.value = label
        l_cell.font = label_font
        l_cell.fill = label_fill
        l_cell.border = cell_border
        l_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        
        # Value cell
        v_cell = ws.cell(row=row_idx, column=2)
        v_cell.value = val
        v_cell.font = value_font
        v_cell.fill = value_fill
        v_cell.border = cell_border
        v_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    row_idx += 1 # Spacer

    # Section Header: Detailed NCCA Fields
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    sec_cell = ws.cell(row=row_idx, column=1)
    sec_cell.value = "NCCA COMPLIANCE REGISTER FIELDS"
    sec_cell.font = section_font
    sec_cell.fill = header_fill
    sec_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_idx].height = 24
    row_idx += 1

    # ---- Dynamic Form Mapping ----
    form_data = profile.form_data or {}
    
    # Create Proxy wrapper helper for easy attribute access
    from modules.heritage.admin_routes import ProxyItem
    item = ProxyItem(profile)

    for field_name, label, field_type, required in config["fields"]:
        val = getattr(item, field_name, None)
        
        # Format values for display
        if val is None:
            val = ""
        elif isinstance(val, (date, datetime)):
            val = val.strftime("%B %d, %Y")
        elif isinstance(val, list):
            val = ", ".join(map(str, val))
        elif isinstance(val, dict):
            # Format lists of informants or similar JSON fields cleanly
            lines = []
            for k, v in val.items():
                lines.append(f"{k}: {v}")
            val = "\n".join(lines)
            
        # Label Cell
        l_cell = ws.cell(row=row_idx, column=1)
        l_cell.value = label
        l_cell.font = label_font
        l_cell.fill = label_fill
        l_cell.border = cell_border
        l_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        
        # Value Cell
        v_cell = ws.cell(row=row_idx, column=2)
        v_cell.value = val
        v_cell.font = value_font
        v_cell.fill = value_fill
        v_cell.border = cell_border
        v_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Dynamic row heights based on content length to prevent clipping
        char_count = len(str(val))
        if char_count > 60:
            row_height = max(20, int(char_count / 3.5))
        else:
            row_height = 22
        ws.row_dimensions[row_idx].height = min(row_height, 120)
        
        row_idx += 1

    # Set perfect column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 65

    # Write to memory stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
